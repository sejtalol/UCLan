# pip install --user python-docx
# pip install --user openpyxl

import openpyxl
import docx
wb= openpyxl.load_workbook('user_new.xlsx')
sheet = wb.get_sheet_by_name('Abstracts Information')
doc = docx.Document()
doc.styles['Normal'].font.name = u'Times New Roman'
nrows = 169
ncols = 5
table = doc.add_table(rows=nrows, cols=ncols, style='Table Grid')
for a in range(1,nrows+1):
    for b in range(1,ncols+1):
        fill = sheet.cell(row=a,column=b).value
#       print('a=',a,'b=',b,'value=',fill)
        if(fill):
           table.cell(a-1,b-1).text = fill
        else:
           table.cell(a-1,b-1).text = u''
doc.save('users.docx')
